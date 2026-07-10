import io
import os
import re
import uuid
import hashlib
from datetime import datetime, date, time
from typing import List

# --- BIBLIOTHEQUES WEB ---
from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, Response, RedirectResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from supabase import create_client, Client

# --- BIBLIOTHEQUES ANALYSE EXCEL ---
import pandas as pd
import pytz
from dateutil import parser as dtparser
from openpyxl import load_workbook

# ==========================================
# 1. CONFIGURATION
# ==========================================

def log_msg(msg: str):
    print(f"[LOG VERCEL] {msg}")

SUPABASE_URL = "https://dlxqgelylxcakrmbkyun.supabase.co"
SUPABASE_KEY = "sb_publishable_mgjSTslsZ_ObnIRxCL10AQ_ix5NSBpz"

try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except Exception as e:
    log_msg(f"Error initializing Supabase: {e}")

app = FastAPI()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
templates = Jinja2Templates(directory="templates")

app.add_middleware(
    SessionMiddleware,
    secret_key="SECRET_SESSION_A_CHANGER_POUR_PROD",
    same_site="lax",
    https_only=True
)

PARIS_TZ = pytz.timezone("Europe/Paris")

# ==========================================
# 2. PARSING EDT
# ==========================================

def normalize_group_label(x):
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass
    s = str(x).strip()
    if not s:
        return None
    m = re.search(r'G\s*\.?\s*(\d+)', s, re.I)
    if m:
        return f'G {m.group(1)}'
    m2 = re.search(r'^(?:groupe)?\s*(\d+)$', s, re.I)
    if m2:
        return f'G {m2.group(1)}'
    return s


def is_time_like(x):
    if x is None:
        return False
    if isinstance(x, (pd.Timestamp, datetime, time)):
        return True
    s = str(x).strip()
    if not s:
        return False
    if re.match(r'^\d{1,2}[:hH]\d{2}(\s*[AaPp][Mm]\.?)*$', s):
        return True
    return False


def to_time(x):
    if x is None:
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, pd.Timestamp):
        return x.to_pydatetime().time()
    if isinstance(x, datetime):
        return x.time()
    s = str(x).strip()
    if not s:
        return None
    s2 = s.replace('h', ':').replace('H', ':')
    try:
        dt = dtparser.parse(s2, dayfirst=True)
        return dt.time()
    except Exception:
        return None


def to_date(x):
    if x is None:
        return None
    if isinstance(x, pd.Timestamp):
        return x.to_pydatetime().date()
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    if not s:
        return None
    try:
        dt = dtparser.parse(s, dayfirst=True, fuzzy=True)
        return dt.date()
    except Exception:
        return None


def get_merged_map(xls_fileobj, sheet_name):
    wb = load_workbook(xls_fileobj, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        r1, r2 = merged.min_row, merged.max_row
        c1, c2 = merged.min_col, merged.max_col
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                merged_map[(r - 1, c - 1)] = (r1 - 1, c1 - 1, r2 - 1, c2 - 1)
    return merged_map


def find_week_rows(df):
    result = []
    for i in range(len(df)):
        val = df.iat[i, 0]
        if val is None:
            continue
        if isinstance(val, str) and re.match(r'^\s*S\s*\.?\s*\d+', val.strip(), re.I):
            result.append(i)
            continue
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            try:
                n = int(val)
                if 1 <= n <= 53 and float(val) == n:
                    result.append(i)
            except (ValueError, TypeError):
                pass
    return result


def find_slot_rows(df):
    return [
        i for i in range(len(df))
        if isinstance(df.iat[i, 0], str)
        and re.match(r'^\s*H\s*\d+', df.iat[i, 0].strip(), re.I)
    ]


def parse_sheet_to_events_json(file_content: bytes, sheet_name: str) -> List[dict]:
    file_io = io.BytesIO(file_content)
    try:
        df = pd.read_excel(file_io, sheet_name=sheet_name, header=None)
    except ValueError:
        log_msg(f"Feuille {sheet_name} introuvable.")
        return []
    except Exception as e:
        log_msg(f"Erreur lecture Excel: {e}")
        return []

    file_io.seek(0)
    merged_map = get_merged_map(file_io, sheet_name)

    nrows, ncols = df.shape
    s_rows = find_week_rows(df)
    h_rows = find_slot_rows(df)

    log_msg(f"[{sheet_name}] Semaines trouvées: {len(s_rows)}, Créneaux trouvés: {len(h_rows)}")

    raw_events = []

    for r in h_rows:
        p_candidates = [s for s in s_rows if s <= r]
        if not p_candidates:
            continue
        p = max(p_candidates)
        date_row = p + 1
        group_row = p + 2

        date_cols = [
            c for c in range(ncols)
            if date_row < nrows and to_date(df.iat[date_row, c]) is not None
        ]

        for c in date_cols:
            for col in (c, c + 1):
                if col >= ncols:
                    continue

                summary = df.iat[r, col]
                if pd.isna(summary) or summary is None:
                    continue
                summary_str = str(summary).strip()
                if not summary_str:
                    continue

                teachers = []
                if (r + 2) < nrows:
                    for off in range(2, 6):
                        idx = r + off
                        if idx >= nrows:
                            break
                        t = df.iat[idx, col]
                        if t is None or pd.isna(t):
                            continue
                        s = str(t).strip()
                        if not s:
                            continue
                        if not is_time_like(s) and to_date(s) is None:
                            teachers.append(s)
                teachers = list(dict.fromkeys(teachers))

                stop_idx = None
                for off in range(1, 12):
                    idx = r + off
                    if idx >= nrows:
                        break
                    if is_time_like(df.iat[idx, col]):
                        stop_idx = idx
                        break
                if stop_idx is None:
                    stop_idx = min(r + 7, nrows)

                desc_parts = []
                for idx in range(r + 1, stop_idx):
                    if idx >= nrows:
                        break
                    cell = df.iat[idx, col]
                    if pd.isna(cell) or cell is None:
                        continue
                    s = str(cell).strip()
                    if not s:
                        continue
                    if to_date(cell) is not None:
                        continue
                    if s in teachers or s == summary_str:
                        continue
                    desc_parts.append(s)
                desc_text = " | ".join(dict.fromkeys(desc_parts))

                start_val, end_val = None, None
                for off in range(1, 13):
                    idx = r + off
                    if idx >= nrows:
                        break
                    v = df.iat[idx, col]
                    if is_time_like(v):
                        if start_val is None:
                            start_val = v
                        elif end_val is None and v != start_val:
                            end_val = v
                            break

                if start_val is None or end_val is None:
                    continue

                start_t, end_t = to_time(start_val), to_time(end_val)
                if start_t is None or end_t is None:
                    continue

                d = to_date(df.iat[date_row, c])
                if d is None:
                    continue
                dtstart = datetime.combine(d, start_t)
                dtend = datetime.combine(d, end_t)

                gl = normalize_group_label(df.iat[group_row, col] if group_row < nrows else None)
                gl_next = normalize_group_label(df.iat[group_row, col + 1] if (col + 1) < ncols else None)
                is_left_col = (col == c)

                groups = set()
                if is_left_col:
                    merged = merged_map.get((r, col))
                    if merged and (r, col + 1) in merged_map:
                        if gl:
                            groups.add(gl)
                        if gl_next:
                            groups.add(gl_next)
                    else:
                        if gl:
                            groups.add(gl)
                else:
                    if gl:
                        groups.add(gl)

                raw_events.append({
                    'summary': summary_str,
                    'teachers': set(teachers),
                    'descriptions': set([desc_text]) if desc_text else set(),
                    'start': dtstart,
                    'end': dtend,
                    'groups': groups
                })

    merged = {}
    for e in raw_events:
        key = (e['summary'], e['start'], e['end'])
        if key not in merged:
            merged[key] = {
                'summary': e['summary'],
                'teachers': set(),
                'descriptions': set(),
                'start': e['start'],
                'end': e['end'],
                'groups': set()
            }
        merged[key]['teachers'].update(e.get('teachers', set()))
        merged[key]['descriptions'].update(e.get('descriptions', set()))
        merged[key]['groups'].update(e.get('groups', set()))

    final_list = []
    for v in merged.values():
        final_list.append({
            'summary': v['summary'],
            'teachers': sorted(list(v['teachers'])),
            'description': " | ".join(sorted(list(v['descriptions']))) if v['descriptions'] else "",
            'start': v['start'].isoformat(),
            'end': v['end'].isoformat(),
            'groups': sorted(list(v['groups']))
        })

    return final_list


# ==========================================
# 2b. PARSING MAQUETTE
# ==========================================

def parse_maquette_sheet(file_content: bytes) -> List[dict]:
    """
    Parse la feuille 'Maquette' du fichier Excel.
    Structure attendue (0-indexé) :
      col 2  = Matière
      col 3-7 = Enseignants 1-5
      col 8  = CM/TD
      col 9  = TP
      col 10 = Autonomie
      col 11 = Examen
      col 12 = Total
      col 13 = Commentaires
      col 14 = Coefficient
    Ligne 0 = groupe de colonnes (headers niveau 1)
    Ligne 1 = headers
    Ligne 2+ = données
    """
    file_io = io.BytesIO(file_content)
    try:
        xls = pd.ExcelFile(file_io)
        maquette_sheet = next((s for s in xls.sheet_names if 'maquette' in s.lower()), None)
        if not maquette_sheet:
            log_msg("Feuille Maquette introuvable.")
            return []
    except Exception as e:
        log_msg(f"Erreur lecture feuilles maquette: {e}")
        return []

    file_io.seek(0)
    try:
        df = pd.read_excel(file_io, sheet_name=maquette_sheet, header=None)
    except Exception as e:
        log_msg(f"Erreur lecture maquette: {e}")
        return []

    if df.shape[1] < 13:
        log_msg("Maquette: pas assez de colonnes.")
        return []

    IGNORE_SUBJECTS = {
        "erasmus day", "forum international", "période entreprise", "periode entreprise",
        "férié", "ferie", "mission à l'international", "mission a l'international",
        "matière", "matières", "divers"
    }

    rows_out = []
    # Contexte de section (semestre / UE) — propagation vers le bas
    current_semester = None
    current_ue = None

    for i in range(2, len(df)):  # skip les 2 lignes de headers
        row = df.iloc[i]

        # Mise à jour du semestre (col 0) si non vide
        if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip():
            current_semester = str(row.iloc[0]).strip()

        # Mise à jour de l'UE (col 1) si non vide
        if pd.notna(row.iloc[1]) and str(row.iloc[1]).strip():
            current_ue = str(row.iloc[1]).strip()

        # Matière (col 2)
        subj_raw = row.iloc[2] if df.shape[1] > 2 else None
        if subj_raw is None or pd.isna(subj_raw):
            continue
        subj = str(subj_raw).strip()
        if not subj:
            continue
        if subj.lower() in IGNORE_SUBJECTS:
            continue

        # Enseignants (cols 3-7)
        teachers = []
        for tc in range(3, 8):
            if tc >= df.shape[1]:
                break
            t = row.iloc[tc]
            if pd.notna(t) and str(t).strip() and str(t).strip() not in [' ', '']:
                teachers.append(str(t).strip())

        def safe_float(v):
            try:
                f = float(v)
                return f if not pd.isna(f) else None
            except Exception:
                return None

        cm_td    = safe_float(row.iloc[8])  if df.shape[1] > 8  else None
        tp       = safe_float(row.iloc[9])  if df.shape[1] > 9  else None
        autonomie= safe_float(row.iloc[10]) if df.shape[1] > 10 else None
        examen   = safe_float(row.iloc[11]) if df.shape[1] > 11 else None
        total    = safe_float(row.iloc[12]) if df.shape[1] > 12 else None
        comment  = str(row.iloc[13]).strip() if df.shape[1] > 13 and pd.notna(row.iloc[13]) else ""
        coeff    = safe_float(row.iloc[14]) if df.shape[1] > 14 else None

        rows_out.append({
            "subject":    subj,
            "semester":   current_semester,
            "ue":         current_ue,
            "teachers":   teachers,
            "cm_td":      cm_td,
            "tp":         tp,
            "autonomie":  autonomie,
            "examen":     examen,
            "total":      total,
            "comment":    comment,
            "coeff":      coeff,
        })

    log_msg(f"Maquette: {len(rows_out)} matières extraites.")
    return rows_out


# ==========================================
# 2c. PARSING ENSEIGNANTS (tarifs)
# ==========================================

def parse_teachers_sheet(file_content: bytes) -> List[dict]:
    """
    Parse la feuille 'Enseignants' du fichier Excel.
    Structure (0-indexé) :
      col 1 = Nom (format "NOM, Prénom" — identique au format utilisé dans EDT P1/P2)
      col 2 = Organisme
      col 3 = Email
      col 4 = Majoration / pourcentage selon le classeur
      col 5 = Tarif horaire théorique
      col 6 = Tarif horaire effectif (= théorique si majoration, sinon théorique x1.25)
      col 7 = Région

    On stocke le tarif EFFECTIF (col 6), qui est celui réellement utilisé dans les
    formules de facturation du classeur Excel (VLOOKUP ... colonne 6).
    """
    file_io = io.BytesIO(file_content)
    try:
        xls = pd.ExcelFile(file_io)
        teachers_sheet = next((s for s in xls.sheet_names if 'enseignant' in s.lower()), None)
        if not teachers_sheet:
            log_msg("Feuille Enseignants introuvable.")
            return []
    except Exception as e:
        log_msg(f"Erreur lecture feuilles enseignants: {e}")
        return []

    file_io.seek(0)
    try:
        df = pd.read_excel(file_io, sheet_name=teachers_sheet, header=None)
    except Exception as e:
        log_msg(f"Erreur lecture enseignants: {e}")
        return []

    if df.shape[1] < 7:
        log_msg("Enseignants: pas assez de colonnes.")
        return []

    def safe_float(v):
        try:
            f = float(v)
            return f if not pd.isna(f) else None
        except Exception:
            return None

    rows_out = []
    seen_names = set()
    for i in range(len(df)):
        name_raw = df.iat[i, 1]
        if name_raw is None or pd.isna(name_raw):
            continue
        name = str(name_raw).strip()
        if not name:
            continue
        # En cas de doublon (le fichier peut lister 2x le même enseignant), garder la 1ère occurrence
        if name in seen_names:
            continue
        seen_names.add(name)

        email = str(df.iat[i, 3]).strip() if df.shape[1] > 3 and pd.notna(df.iat[i, 3]) else ""
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            log_msg(f"Email enseignant ignoré (format invalide) pour {name}: {email}")
            email = ""

        rate = safe_float(df.iat[i, 6])  # tarif effectif (colonne G / index 6)
        rows_out.append({
            "name": name,
            "organism": str(df.iat[i, 2]).strip() if pd.notna(df.iat[i, 2]) else "",
            "email": email,
            "hourly_rate": rate if rate is not None else 0.0,
        })

    log_msg(f"Enseignants: {len(rows_out)} tarifs extraits.")
    return rows_out


# ==========================================
# 3. GENERATEUR ICS
# ==========================================

def escape_ical_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace('\\', '\\\\').replace('\n', '\\n').replace(',', '\\,').replace(';', '\\;')
    return s


def build_paris_vtimezone_text():
    return "\n".join([
        "BEGIN:VTIMEZONE",
        "TZID:Europe/Paris",
        "X-LIC-LOCATION:Europe/Paris",
        "BEGIN:DAYLIGHT",
        "TZOFFSETFROM:+0100",
        "TZOFFSETTO:+0200",
        "TZNAME:CEST",
        "DTSTART:19700329T020000",
        "RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU",
        "END:DAYLIGHT",
        "BEGIN:STANDARD",
        "TZOFFSETFROM:+0200",
        "TZOFFSETTO:+0100",
        "TZNAME:CET",
        "DTSTART:19701025T030000",
        "RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU",
        "END:STANDARD",
        "END:VTIMEZONE"
    ])


def events_to_ics_string(events: List[dict], tzname='Europe/Paris', uid_namespace: str = 'edt') -> str:
    tz = pytz.timezone(tzname)

    header = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//EDT Export//FR',
        'CALSCALE:GREGORIAN',
    ]

    body = [build_paris_vtimezone_text()]

    for ev in events:
        uid_source = '|'.join([uid_namespace, str(ev.get('start','')), str(ev.get('end','')), str(ev.get('summary','')), ','.join(ev.get('groups',[]) or [])])
        uid = hashlib.sha256(uid_source.encode('utf-8')).hexdigest() + '@cesi-edt'

        try:
            start_dt = datetime.fromisoformat(ev['start'])
            end_dt = datetime.fromisoformat(ev['end'])
        except ValueError:
            continue

        if start_dt.tzinfo is None:
            start_loc = tz.localize(start_dt)
        else:
            start_loc = start_dt.astimezone(tz)

        if end_dt.tzinfo is None:
            end_loc = tz.localize(end_dt)
        else:
            end_loc = end_dt.astimezone(tz)

        dtstart = start_loc.strftime('%Y%m%dT%H%M%S')
        dtend = end_loc.strftime('%Y%m%dT%H%M%S')

        # Pour les flux multi-promos (notamment l'abonnement enseignant),
        # rendre la promo et les groupes visibles directement dans le titre.
        promo_label = str(ev.get('promo_label') or '').strip().upper()
        groups = [str(g).strip() for g in (ev.get('groups') or []) if str(g).strip()]
        title_parts = []
        if promo_label:
            title_parts.append(promo_label)
        title_parts.extend(groups)
        raw_summary = str(ev.get('summary') or '')
        if title_parts:
            raw_summary = f"[{' · '.join(title_parts)}] {raw_summary}"
        summary = escape_ical_text(raw_summary)

        desc_lines = []
        if promo_label:
            desc_lines.append('Promotion : ' + promo_label)
        if ev.get('description'):
            desc_lines.append(ev['description'])

        teachers = ev.get('teachers', [])
        if teachers:
            desc_lines.append('Enseignant(s): ' + ' / '.join(teachers))

        if groups:
            if len(groups) == 1:
                desc_lines.append('Groupe : ' + groups[0])
            else:
                desc_lines.append('Groupes : ' + ' et '.join(groups))

        room = str(ev.get('room') or ev.get('location') or '').strip()
        if room:
            desc_lines.append('Salle : ' + room)

        description = escape_ical_text('\n'.join(desc_lines))

        event_lines = [
            'BEGIN:VEVENT',
            f'UID:{uid}',
            f'DTSTAMP:{datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")}',
            f'DTSTART;TZID={tzname}:{dtstart}',
            f'DTEND;TZID={tzname}:{dtend}',
            f'SUMMARY:{summary}',
        ]
        if room:
            event_lines.append(f'LOCATION:{escape_ical_text(room)}')
        event_lines.extend([
            f'DESCRIPTION:{description}',
            'END:VEVENT'
        ])
        body.extend(event_lines)

    footer = ['END:VCALENDAR']
    return '\n'.join(header + body + footer)


# ==========================================
# 4. ROUTES & AUTH
# ==========================================

def get_current_user(request: Request):
    return request.session.get("user")


def hash_password(password: str):
    return hashlib.sha256(password.encode()).hexdigest()


def convert_updated_at(plannings: list) -> list:
    for p in plannings:
        if p.get("updated_at"):
            try:
                dt = datetime.fromisoformat(p["updated_at"].replace("Z", "+00:00"))
                dt_paris = dt.astimezone(PARIS_TZ)
                p["updated_at"] = dt_paris.isoformat()
            except Exception:
                pass
    return plannings


@app.get("/", response_class=HTMLResponse)
async def home(request: Request, filter: str = "my"):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    query = supabase.table("plannings").select("slug, name, year, updated_at, creator")

    if filter == "my":
        query = query.eq("creator", user)

    response = query.execute()
    plannings = convert_updated_at(response.data)

    return templates.TemplateResponse(request, "index.html", {
        "user": user,
        "plannings": plannings,
        "current_filter": filter
    })


# --- LOGIN / REGISTER ---

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", {"mode": "login"})


@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    return templates.TemplateResponse(request, "login.html", {"mode": "register"})


@app.post("/login")
async def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    hashed = hash_password(password)
    res = supabase.table("users").select("*").eq("username", username).eq("password_hash", hashed).execute()

    if len(res.data) > 0:
        request.session["user"] = username
        return RedirectResponse(url="/", status_code=303)
    else:
        return templates.TemplateResponse(request, "login.html", {
            "mode": "login",
            "error": "Identifiants incorrects"
        })


@app.post("/register")
async def register_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    verification: str = Form(...)
):
    if verification.strip().upper() != "EUROVISION":
        return templates.TemplateResponse(request, "login.html", {
            "mode": "register",
            "error": "Code de vérification incorrect !"
        })

    check = supabase.table("users").select("*").eq("username", username).execute()
    if len(check.data) > 0:
        return templates.TemplateResponse(request, "login.html", {
            "mode": "register",
            "error": "Ce pseudo est déjà pris."
        })

    hashed = hash_password(password)
    try:
        supabase.table("users").insert({"username": username, "password_hash": hashed}).execute()
        request.session["user"] = username
        return RedirectResponse(url="/", status_code=303)
    except Exception as e:
        log_msg(f"Erreur inscription: {e}")
        return templates.TemplateResponse(request, "login.html", {
            "mode": "register",
            "error": "Erreur technique lors de l'inscription."
        })


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


# --- ACTIONS CALENDRIER ---

@app.post("/create")
async def create_calendar(request: Request, promo_name: str = Form(...), school_year: str = Form(...)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")

    slug = f"{promo_name}-{school_year}".lower().replace(" ", "-")
    data = {"slug": slug, "name": promo_name, "year": school_year, "creator": user}
    try:
        supabase.table("plannings").insert(data).execute()
        log_msg(f"Nouveau planning créé: {slug} par {user}")
    except Exception as e:
        log_msg(f"Erreur création (existe probablement déjà): {e}")

    return RedirectResponse("/", status_code=303)


@app.post("/upload/{slug}")
async def upload_excel(slug: str, request: Request, file: UploadFile = File(...)):
    if not get_current_user(request):
        return RedirectResponse("/login")

    log_msg(f"Début upload pour {slug}")
    content = await file.read()

    try:
        xls = pd.ExcelFile(io.BytesIO(content))
        log_msg(f"Feuilles disponibles: {xls.sheet_names}")
    except Exception as e:
        log_msg(f"Erreur lecture feuilles: {e}")

    events_p1 = parse_sheet_to_events_json(content, "EDT P1")
    events_p2 = parse_sheet_to_events_json(content, "EDT P2")
    maquette_data = parse_maquette_sheet(content)
    teachers_data = parse_teachers_sheet(content)

    log_msg(f"Events trouvés - P1: {len(events_p1)}, P2: {len(events_p2)}, Maquette: {len(maquette_data)}, Enseignants: {len(teachers_data)}")

    try:
        supabase.table("plannings").update({
            "events_p1": events_p1,
            "events_p2": events_p2,
            "maquette_data": maquette_data,
            "teachers_data": teachers_data,
            "updated_at": datetime.now(PARIS_TZ).isoformat()
        }).eq("slug", slug).execute()
    except Exception as e:
        log_msg(f"Erreur mise à jour BDD: {e}")

    return RedirectResponse("/", status_code=303)


# --- URL PUBLIQUE POUR OUTLOOK ---

@app.get("/ics/{slug}/{group}.ics")
async def get_ics_file(slug: str, group: str):
    group_clean = group.upper()
    if group_clean not in ["P1", "P2"]:
        raise HTTPException(404, detail="Groupe inconnu (utiliser P1 ou P2)")

    try:
        res = supabase.table("plannings").select(f"events_{group_clean.lower()}").ilike("slug", slug).execute()

        if not res.data:
            raise HTTPException(404, detail="Planning introuvable")

        events_json = res.data[0].get(f"events_{group_clean.lower()}", []) or []
        ics_content = events_to_ics_string(events_json)
        filename = f"{slug}_{group_clean}.ics"

        return Response(content=ics_content, media_type="text/calendar", headers={
            "Content-Disposition": f"attachment; filename={filename}"
        })

    except HTTPException:
        raise
    except Exception as e:
        log_msg(f"Erreur ICS generation: {e}")
        raise HTTPException(500, detail="Erreur interne")


@app.get("/ics/{slug}/enseignant.ics")
async def get_teacher_ics_file(slug: str, teacher: str):
    """Flux ICS public et stable d'un enseignant, utilisable en abonnement calendrier."""
    try:
        res = supabase.table("plannings").select("events_p1, events_p2, teachers_data").ilike("slug", slug).execute()
        if not res.data:
            raise HTTPException(404, detail="Planning introuvable")

        data = res.data[0]
        known_teachers = {str(t.get("name", "")).strip() for t in (data.get("teachers_data") or [])}
        # Conserver l'origine P1/P2 : le calendrier de l'enseignant mélange les
        # deux promotions et l'information doit donc voyager avec chaque séance.
        all_events = []
        for ev in (data.get("events_p1") or []):
            all_events.append({**ev, "promo_label": "P1"})
        for ev in (data.get("events_p2") or []):
            all_events.append({**ev, "promo_label": "P2"})
        events = [ev for ev in all_events if teacher in (ev.get("teachers") or [])]
        if teacher not in known_teachers and not events:
            raise HTTPException(404, detail="Enseignant introuvable")

        ics_content = events_to_ics_string(events, uid_namespace=f"{slug}:teacher:{teacher}")
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", teacher).strip("_") or "enseignant"
        return Response(content=ics_content, media_type="text/calendar; charset=utf-8", headers={
            "Content-Disposition": f'inline; filename="Planning_{safe_name}.ics"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
        })
    except HTTPException:
        raise
    except Exception as e:
        log_msg(f"Erreur ICS enseignant: {e}")
        raise HTTPException(500, detail="Erreur interne")


# --- VUE PUBLIQUE CALENDRIER ---

@app.get("/calendrier/{slug}", response_class=HTMLResponse)
async def view_calendar(slug: str, request: Request):
    res = supabase.table("plannings").select("slug, name, year").ilike("slug", slug).execute()
    if not res.data:
        raise HTTPException(404, detail="Planning introuvable")
    p = res.data[0]
    return templates.TemplateResponse(request, "calendar_view.html", {
        "slug": p["slug"],
        "planning_name": p["name"],
        "planning_year": p.get("year", ""),
    })


# --- VUE TABLEAU DE BORD (protégée) ---

@app.get("/dashboard/{slug}", response_class=HTMLResponse)
async def view_dashboard(slug: str, request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    res = supabase.table("plannings").select("slug, name, year").ilike("slug", slug).execute()
    if not res.data:
        raise HTTPException(404, detail="Planning introuvable")
    p = res.data[0]
    return templates.TemplateResponse(request, "dashboard.html", {
        "slug": p["slug"],
        "planning_name": p["name"],
        "planning_year": p.get("year", ""),
        "user": user,
    })


# --- APIS JSON ---

@app.get("/api/events/{slug}/{group}")
async def api_events(slug: str, group: str):
    group_clean = group.lower()
    if group_clean not in ["p1", "p2"]:
        raise HTTPException(404)
    try:
        res = supabase.table("plannings").select(f"events_{group_clean}").ilike("slug", slug).execute()
        if not res.data:
            raise HTTPException(404)
        return res.data[0].get(f"events_{group_clean}", []) or []
    except HTTPException:
        raise
    except Exception as e:
        log_msg(f"Erreur API events: {e}")
        raise HTTPException(500, detail="Erreur interne")


@app.get("/api/maquette/{slug}")
async def api_maquette(slug: str):
    """Retourne les données de la maquette pédagogique."""
    try:
        res = supabase.table("plannings").select("maquette_data").ilike("slug", slug).execute()
        if not res.data:
            raise HTTPException(404)
        return res.data[0].get("maquette_data", []) or []
    except HTTPException:
        raise
    except Exception as e:
        log_msg(f"Erreur API maquette: {e}")
        raise HTTPException(500, detail="Erreur interne")


@app.get("/api/dashboard-data/{slug}")
async def api_dashboard_data(slug: str):
    """Retourne en une seule requête : events_p1, events_p2, maquette_data, teachers_data."""
    try:
        res = supabase.table("plannings").select(
            "events_p1, events_p2, maquette_data, teachers_data"
        ).ilike("slug", slug).execute()
        if not res.data:
            raise HTTPException(404)
        d = res.data[0]
        return {
            "events_p1":     d.get("events_p1", []) or [],
            "events_p2":     d.get("events_p2", []) or [],
            "maquette_data": d.get("maquette_data", []) or [],
            "teachers_data": d.get("teachers_data", []) or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        log_msg(f"Erreur API dashboard-data: {e}")
        raise HTTPException(500, detail="Erreur interne")
